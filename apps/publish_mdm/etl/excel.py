import re
import structlog
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from typing import Generator

logger = structlog.getLogger(__name__)


def get_header(sheet: Worksheet, column_name: str) -> Cell:
    """Find the column header cell by name in the first row of the sheet."""
    header_cell = None
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            if cell.value == column_name:
                header_cell = cell
                break
    if not header_cell:
        logger.debug("Could not find column header", column_name=column_name)
    return header_cell


def get_column_cell_by_value(column_header: Cell, value: str, is_regex=False) -> Cell | None:
    """Find the cell by searching down the column for the value."""
    sheet: Worksheet = column_header.parent
    target_cell = None
    for row in sheet.iter_rows(
        min_row=2,
        max_row=sheet.max_row,
        min_col=column_header.column,
        max_col=column_header.column,
    ):
        for cell in row:
            if is_regex:
                if isinstance(cell.value, (str, bytes)) and re.match(value, cell.value):
                    target_cell = cell
                    break
            elif cell.value == value:
                target_cell = cell
                break
    if not target_cell:
        logger.debug(
            "Could not find value in column",
            column_name=column_header.value,
            value=value,
            is_regex=is_regex,
        )
    return target_cell


def find_cells_containing_value(sheet: Worksheet, value: str) -> Generator[Cell, None, None]:
    """Find every cell that contains the value."""
    # skip header row
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if cell.value is not None and value in str(cell.value):
                yield cell
