import structlog
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel

logger = structlog.getLogger(__name__)


class TemplateVariable(BaseModel):
    name: str
    value: str | int | float


def get_header(sheet: Worksheet, column_name: str) -> Cell:
    """Find the column header cell by name in the first row of the sheet."""
    header_cell = None
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            if cell.value == column_name:
                header_cell = cell
                break
    if not header_cell:
        logger.warning("Could not find column header", column_name=column_name)
    return header_cell


def get_cell_by_value(column_header: Cell, value: str) -> Cell:
    """Find the cell by searching down the column for the value."""
    sheet = column_header.parent
    target_cell = None
    for row in sheet.iter_rows(
        min_row=2,
        max_row=sheet.max_row,
        min_col=column_header.column,
        max_col=column_header.column,
    ):
        for cell in row:
            if cell.value == value:
                target_cell = cell
                break
    if not target_cell:
        logger.warning(
            "Could not find value in column", column_name=column_header.value, value=value
        )

    return target_cell


def set_template_variables(sheet: Worksheet, variables: list[TemplateVariable]):
    """Fill in the template variables on the survey sheet.

    Variables are just `calculate` rows in the survey sheet, so we need to find
    the variable in the `name` column and then offset to the `calculation`
    column to fill in the value.
    """
    name_header = get_header(sheet=sheet, column_name="name")
    calculation_column = get_header(sheet=sheet, column_name="calculation").column
    # Calculate the number of columns over to the calculation column
    offset = calculation_column - name_header.column
    for variable in variables:
        variable_cell = get_cell_by_value(column_header=name_header, value=variable.name)
        variable_cell.offset(column=offset).value = variable.value
