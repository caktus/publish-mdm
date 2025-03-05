import hashlib
import re
from enum import StrEnum

import structlog
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, computed_field

from .excel import find_cells_containing_value, get_column_cell_by_value, get_header

logger = structlog.getLogger(__name__)


class VariableTransform(StrEnum):
    SHA256_DIGEST = "sha256-digest"

    @classmethod
    def choices(cls):
        return [(item.value, item.name) for item in cls]


class TemplateVariable(BaseModel):
    """A template variable to fill in the survey sheet."""

    name: str
    transform: VariableTransform | None = None
    value: str

    @computed_field
    @property
    def rendered_value(self) -> str:
        """XForm expressions must be quoted. Otherwise, an error is raised like:
        Invalid calculate for the bind / null in expression
        """
        value = self.value
        if self.transform == VariableTransform.SHA256_DIGEST:
            value = hashlib.sha256(value.encode()).hexdigest()
            logger.debug(
                "Transformed variable value",
                name=self.name,
                value=value,
                transform=self.transform.value,
            )
        return f'"{value}"'


def set_survey_template_variables(sheet: Worksheet, variables: list[TemplateVariable]):
    """Fill in the template variables on the survey sheet.

    Variables are just `calculate` rows in the survey sheet, so we need to find
    the variable in the `name` column and then offset to the `calculation`
    column to fill in the value.
    """
    name_header = get_header(sheet=sheet, column_name="name")
    calculation_column = get_header(sheet=sheet, column_name="calculation").column
    # Calculate the number of columns over to the calculation column
    offset = calculation_column - name_header.column
    for variable in variables:
        variable_cell = get_column_cell_by_value(column_header=name_header, value=variable.name)
        calculation_cell = variable_cell.offset(column=offset)
        value = variable.rendered_value
        logger.debug(
            "Setting variable value",
            variable=variable.name,
            value=value,
            cell=calculation_cell.coordinate,
        )
        calculation_cell.value = value


def set_survey_attachments(sheet: Worksheet, attachments: dict | None = None):
    """Fill in the static attachments on the survey sheet.

    Attachment columns' headers are either "media::image", "media::audio", or "media::video".
    The `attachments` dict should contain data from the ProjectAttachment model,
    with the values from the `name` and `file` fields as the keys and values respectively.
    The dict will be updated in place to remove attachments that are not detected in the
    survey sheet. To detect an attachment, we will look for the key from the `attachments`
    dict in the "media::*" columns.
    """
    if not attachments:
        return
    # Get all the "media::" columns in the sheet
    media_headers = []
    for media_type in ("image", "audio", "video"):
        header = get_header(sheet=sheet, column_name=f"media::{media_type}")
        if header:
            media_headers.append(header)
    if not media_headers:
        # No media headers found, so there are no attachments in the form
        attachments.clear()
        return
    logger.debug("Found media headers", media_headers=media_headers)
    for name, file in list(attachments.items()):
        found_attachment_name = False
        for media_header in media_headers:
            attachment_cell = get_column_cell_by_value(column_header=media_header, value=name)
            if attachment_cell:
                found_attachment_name = True
                logger.debug(
                    "Found attachment reference",
                    name=name,
                    cell=attachment_cell.coordinate,
                    header=media_header.value,
                )
                break
        if not found_attachment_name:
            # This attachment is not being used in this form. Remove it from the dict
            del attachments[name]
            continue
    if attachments:
        # Log the attachments that are being used in the form
        logger.debug("All found form attachments", attachments=list(attachments.keys()))


def update_setting_variables(
    sheet: Worksheet, title_base: str, form_id_base: str, app_user: str, version: str
):
    """Update the settings sheet to be specific to the app user.

    Available settings: https://docs.getodk.org/xlsform/#the-settings-sheet
    """
    # Append [<app_user>] to the form_title
    # The values are in the 2nd row of the sheet, so we offset by 1 for each settting
    form_title_cell = get_header(sheet=sheet, column_name="form_title").offset(row=1)
    form_title_cell.value = f"{title_base} [{app_user}]"
    logger.debug("Set form_title", cell=form_title_cell.coordinate, value=form_title_cell.value)
    # Append _<app_user> to the form_id
    form_id_cell = get_header(sheet=sheet, column_name="form_id").offset(row=1)
    form_id_cell.value = f"{form_id_base}_{app_user}"
    logger.debug("Set form_id", cell=form_id_cell.coordinate, value=form_id_cell.value)
    # Update the version, with the current date in YYYY-MM-DD format
    version_cell = get_header(sheet=sheet, column_name="version").offset(row=1)
    version_cell.value = version
    logger.debug("Set version", cell=version_cell.coordinate, value=version_cell.value)


def discover_entity_lists(workbook: Workbook) -> list[str]:
    """Discover the entity lists in the survey sheet."""
    patterns = (
        # instance('fruits')/root/...
        re.compile(r"instance\('([\w_]+)'"),
        # pulldata('fruits', '')...
        re.compile(r"pulldata\('([\w_]+)'"),
        # select_one_from_file fruits.csv
        re.compile(r"([\w_]+).csv$"),
    )
    entity_lists = set()
    # Find entity lists in the `survey` sheet
    for row in workbook["survey"].iter_rows(min_row=2):
        for cell in row:
            if cell.value:
                for pattern in patterns:
                    match = pattern.search(str(cell.value))
                    if match:
                        entity_lists.add(match.group(1))
                        logger.debug(
                            "Discovered entity list",
                            entity_list=match.group(1),
                            cell=cell.coordinate,
                            sheet="survey",
                        )
    # Find entity lists in the `entities` sheet
    if "entities" in workbook:
        for row in workbook["entities"].iter_rows(min_row=2, min_col=1, max_col=1):
            for cell in row:
                if cell.value:
                    entity_lists.add(str(cell.value))
                    logger.debug(
                        "Discovered entity list",
                        entity_list=cell.value,
                        cell=cell.coordinate,
                        sheet="entities",
                    )
    logger.debug("Discovered entity lists", entity_lists=entity_lists)
    return entity_lists


def build_entity_list_mapping(workbook: Workbook, app_user: str) -> dict[str, str]:
    """Build a mapping of app user entity lists to new entity list names."""
    entity_lists = discover_entity_lists(workbook=workbook)
    substitutes = {}
    for entity_list in entity_lists:
        if entity_list.endswith("_APP_USER"):
            name_substitute = f"{entity_list.rsplit('_APP_USER', 1)[0]}_{app_user}"
            substitutes[entity_list] = name_substitute
    return substitutes


def update_entity_references(workbook: Workbook, entity_list_mapping: dict[str, str] = None):
    """Update references to entity lists in the workbook, based on the mapping.

    For example: `filename_APP_USER.csv` -> `filename_11030.csv`
    """
    if not entity_list_mapping:
        entity_list_mapping = {}

    # Update references to entity lists in the `survey`` sheet
    for entity_list_orig, entity_list_new in entity_list_mapping.items():
        # Replace references to "filename_APP_USER.csv"
        for cell in find_cells_containing_value(
            sheet=workbook["survey"], value=f"{entity_list_orig}.csv"
        ):
            cell.value = cell.value.replace(f"{entity_list_orig}.csv", f"{entity_list_new}.csv")
            logger.debug(
                "Updated entity list reference",
                cell=cell.coordinate,
                value=cell.value,
                sheet="survey",
            )
        # Replace references to the fill in pulldata() and instance()
        # (in single quotes, e.g., "'filename_APP_USER'")
        for cell in find_cells_containing_value(
            sheet=workbook["survey"], value=f"'{entity_list_orig}'"
        ):
            cell.value = cell.value.replace(f"'{entity_list_orig}'", f"'{entity_list_new}'")
            logger.debug(
                "Updated entity list pulldata reference",
                cell=cell.coordinate,
                value=cell.value,
                sheet="survey",
            )

    # Update references to entity lists in the `entities` sheet
    if "entities" in workbook:
        # Skip the header row
        for row in workbook["entities"].iter_rows(min_row=2):
            for cell in row:
                if cell.value in entity_list_mapping:
                    cell.value = entity_list_mapping[cell.value]
                    logger.debug(
                        "Updated entity list reference",
                        cell=cell.coordinate,
                        value=cell.value,
                        sheet="entities",
                    )
